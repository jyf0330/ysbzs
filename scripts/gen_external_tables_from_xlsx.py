#!/usr/bin/env python3
"""Generate split JSON artifacts and runtime tables from the external xlsx table.

Input:
  docs/01_游戏设计（策划主导）/关卡策划/ysbzs_英雄表_最新设定_20260531.xlsx
Output:
    external-data/unit_patches.json
    external-data/shop_unlock_pools.json
    external-data/id_alias.json
    external-data/meta.json
    external-tables.js
"""

from __future__ import annotations

import datetime as _dt
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "docs/01_游戏设计（策划主导）/关卡策划/ysbzs_英雄表_最新设定_20260531.xlsx"
OUT = ROOT / "external-tables.js"
OUT_DIR = ROOT / "external-data"

# xlsx id/name -> runtime id
ID_ALIAS = {
    "flame_sprite": "fire_starter",
    "drop_sprite": "water_droplet",
    "bubble_sprite": "bubble_sprite",
    "sprout_summoner": "sprout_summoner",
    "fluff_speaker": "fluff_speaker",
    "boom_sprite": "boom_sprite",
    "split_sprite": "split_sprite",
}

NAME_ALIAS = {
    "火苗灵": "fire_starter",
    "滴滴灵": "water_droplet",
    "泡泡灵": "bubble_sprite",
    "召芽灵": "sprout_summoner",
    "绒语灵": "fluff_speaker",
    "爆爆灵": "boom_sprite",
    "分分灵": "split_sprite",
}


def _n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    if s.isdigit():
        return int(s)
    return None


def parse_slot_tiers(text: str):
    if not text:
        return None
    s = str(text)

    # 槽1 火+1；槽2 火+2；槽3 火+3
    vals = [int(x) for x in re.findall(r"槽\d+[^+\d]*\+(\d+)", s)]
    if len(vals) >= 2:
        return vals[:3]

    # 2目标 治疗+1/+2/+3
    m = re.search(r"\+(\d+)\s*/\+?(\d+)\s*/\+?(\d+)", s)
    if m:
        return [int(m.group(1)), int(m.group(2)), int(m.group(3))]

    return None


def parse_shop_pool(raw: str):
    if not raw:
        return []
    s = str(raw).replace("、", ",").replace("，", ",")
    return [x.strip() for x in s.split(",") if x and x.strip()]


def dump_json(path: Path, payload):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"xlsx not found: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["01_英雄总表_当前3槽版"]
    ws_pool = wb["03_商店解锁池"]

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    patches = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue

        x_id = str(r[col["ID"]]).strip() if r[col["ID"]] is not None else ""
        x_name = str(r[col["英雄名"]]).strip() if r[col["英雄名"]] is not None else ""
        rid = ID_ALIAS.get(x_id) or NAME_ALIAS.get(x_name)
        if not rid:
            # only output rows we can map to current runtime units
            continue

        hp1 = _n(r[col["铜HP"]])
        hp2 = _n(r[col["银HP"]])
        hp3 = _n(r[col["金HP"]])

        t1 = parse_slot_tiers(r[col["铜动作（3槽）"]])
        t2 = parse_slot_tiers(r[col["银动作（3槽）"]])
        t3 = parse_slot_tiers(r[col["金动作（3槽）"]])

        levels = {}
        if hp1 is not None or t1:
            levels["1"] = {k: v for k, v in {"hp": hp1, "slotTiers": t1}.items() if v is not None}
        if hp2 is not None or t2:
            levels["2"] = {k: v for k, v in {"hp": hp2, "slotTiers": t2}.items() if v is not None}
        if hp3 is not None or t3:
            levels["3"] = {k: v for k, v in {"hp": hp3, "slotTiers": t3}.items() if v is not None}

        patch = {
            "name": x_name or None,
            "levels": levels,
        }
        patch = {k: v for k, v in patch.items() if v}
        if patch:
            patches[rid] = patch

    pool_rows = []
    pool_headers = [c.value for c in ws_pool[1]]
    pool_col = {h: i for i, h in enumerate(pool_headers)}
    for r in ws_pool.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        day = _n(r[pool_col["解锁日"]])
        if day is None:
            continue
        pool_rows.append(
            {
                "day": day,
                "pool": parse_shop_pool(r[pool_col["商店池"]]),
                "teachingFocus": (str(r[pool_col["建议教学重点"]]).strip() if r[pool_col["建议教学重点"]] else ""),
                "exclude": (str(r[pool_col["不要在此阶段塞入"]]).strip() if r[pool_col["不要在此阶段塞入"]] else ""),
            }
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    unit_patches_payload = {
        "UNIT_PATCHES": patches,
    }
    shop_pool_payload = {
        "SHOP_UNLOCK_POOLS": pool_rows,
    }
    id_alias_payload = {
        "ID_ALIAS": ID_ALIAS,
        "NAME_ALIAS": NAME_ALIAS,
    }
    meta_payload = {
        "source": str(SRC.relative_to(ROOT)),
        "generatedAt": _dt.datetime.now().isoformat(timespec="seconds"),
        "note": "Auto generated from external xlsx hero table",
    }

    dump_json(OUT_DIR / "unit_patches.json", unit_patches_payload)
    dump_json(OUT_DIR / "shop_unlock_pools.json", shop_pool_payload)
    dump_json(OUT_DIR / "id_alias.json", id_alias_payload)
    dump_json(OUT_DIR / "meta.json", meta_payload)

    payload = {
        "UNIT_PATCHES": patches,
        "SHOP_UNLOCK_POOLS": pool_rows,
        "__meta": meta_payload,
    }

    js = (
        "/* Auto-generated file. Do not edit manually. */\n"
        "(function(){\n"
        "  var g = (typeof globalThis !== 'undefined') ? globalThis : window;\n"
        "  var prev = g.__YSBZS_TABLES__ || {};\n"
        "  var next = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n"
        "  g.__YSBZS_TABLES__ = Object.assign({}, prev, next);\n"
        "})();\n"
    )

    OUT.write_text(js, encoding="utf-8")
    print(f"generated: {OUT}")
    print(f"generated dir: {OUT_DIR}")
    print(f"unit patches: {len(patches)}")
    print(f"shop unlock rows: {len(pool_rows)}")


if __name__ == "__main__":
    main()
