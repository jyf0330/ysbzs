"""One-time, idempotent identity-only worksheet addition from the locked DB.

Does not export raw source payload, descriptions or assets. Existing cells are
checked before replacement; CSV output remains the normal master exporter.
"""
import argparse
import hashlib
import json
import pathlib
import sqlite3
import tempfile
import openpyxl
import export_master_to_csv as source
import bazaar_run_source_views as view

NAMES = {'Spices','Pickpocket','Honed Arsenal','Aggressive','Stunning Strike',
         'Venomous Blade','Valley Fever','Second Wind','Frontline Logistics',
         'Parting Shot','Beast Unleashed','Inspired Rage','Bloodhound',
         'Iron Sharpens Iron','Toxin Injector'}

def main():
    parser=argparse.ArgumentParser();parser.add_argument('--db',required=True);parser.add_argument('--master',required=True)
    args=parser.parse_args();db=pathlib.Path(args.db);master=pathlib.Path(args.master)
    with db.open('rb') as f:
        if hashlib.file_digest(f,'sha256').hexdigest()!=source.RELOCKED_LOCAL_CACHE_GAMEDATA_SHA256:
            raise ValueError('RUN_SOURCE_DB_HASH_MISMATCH')
    rows=[];seen=set()
    with sqlite3.connect(db.resolve().as_uri()+'?mode=ro',uri=True) as connection:
        for uid,raw in connection.execute('SELECT Id,Data FROM cards'):
            data=json.loads(raw);name=data.get('InternalName')
            if name not in NAMES:continue
            if name in seen:raise ValueError('RUN_SOURCE_NAME_AMBIGUOUS')
            seen.add(name)
            values=[view.VIEW_ID,data['Type'].lower(),uid,
                    '|'.join(sorted(h.lower() for h in data['Heroes'])),data['SpawningEligibility'],
                    '|'.join(q for q in view.QUALITIES if q.title() in data['Tiers'])]
            rows.append(dict(zip(view.MEMBER_HEADERS,values)))
    if seen!=NAMES or view.member_digest(rows)!=view.MEMBERS_SHA256:
        raise ValueError('RUN_SOURCE_EXTRACTED_IDENTITY_MISMATCH')
    workbook=openpyxl.load_workbook(master)
    additions={view.VIEW_SHEET:(view.VIEW_HEADERS,[view.view_row(source.RELOCKED_LOCAL_CACHE_SNAPSHOT_ID)]),
               view.MEMBER_SHEET:(view.MEMBER_HEADERS,sorted(rows,key=lambda r:(r['source_type'],r['source_uuid'])))}
    before={s.title:list(s.values) for s in workbook if s.title not in additions}
    changed=False
    for name,(headers,records) in additions.items():
        values=[tuple(headers)]+[tuple(r[k] for k in headers) for r in records]
        if name in workbook:
            existing=list(workbook[name].values)
            original_view=[existing[0]]+[row for row in existing[1:] if row[0]==view.VIEW_ID]
            if original_view!=values:raise ValueError('RUN_SOURCE_EXISTING_SHEET_CONFLICT:'+name)
        else:
            changed=True
            sheet=workbook.create_sheet(name)
            for row in values:sheet.append(row)
    if not changed:
        source.generated_reference_source_tables(master)
        workbook.close()
        print('PASS 15 identities already match; no workbook rewrite')
        return
    with tempfile.NamedTemporaryFile(prefix=master.stem+'.run-source.',suffix='.xlsx',dir=master.parent,delete=False) as handle:
        temporary=pathlib.Path(handle.name)
    workbook.save(temporary)
    check=openpyxl.load_workbook(temporary,read_only=True,data_only=False)
    if any(list(check[name].values)!=values for name,values in before.items()):
        raise ValueError('RUN_SOURCE_OLD_SHEET_CHANGED')
    source.generated_reference_source_tables(temporary)
    check.close();temporary.replace(master)
    print('PASS 15 identities; all',len(before),'existing sheet cell values unchanged; reference only')

if __name__=='__main__':main()
