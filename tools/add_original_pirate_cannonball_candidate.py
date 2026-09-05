"""Verify locked local mechanics then create only the isolated candidate workbook."""
import hashlib
import json
from pathlib import Path
import sqlite3
import tempfile
import openpyxl
import export_original_pirate_cannonball_candidate as c


def migrate_legacy_workbook(path, expected):
    """Add the v39 availability column only after the old rows match exactly."""
    book = openpyxl.load_workbook(path)
    try:
        sheet = book['BZ_ITEMS']
        current_headers = [cell.value for cell in sheet[1]]
        target_headers = c.e.DOMAIN_HEADERS['46_bz_items.csv']
        legacy_headers = [header for header in target_headers if header != 'availability']
        if current_headers == target_headers:
            return False
        if current_headers != legacy_headers:
            raise ValueError('CANDIDATE_LEGACY_HEADERS_INVALID')
        expected_rows = [
            {key: value for key, value in row.items() if key != 'availability'}
            for row in expected['46_bz_items.csv']
        ]
        actual_rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            actual_rows.append(dict(zip(legacy_headers, ['' if value is None else str(value) for value in values])))
        if sorted(actual_rows, key=lambda row: tuple(row[key] for key in legacy_headers)) != sorted(expected_rows, key=lambda row: tuple(row[key] for key in legacy_headers)):
            raise ValueError('CANDIDATE_LEGACY_ROWS_INVALID')
        column = target_headers.index('availability') + 1
        sheet.insert_cols(column)
        sheet.cell(row=1, column=column, value='availability')
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column, value='run_acquirable')
        with tempfile.NamedTemporaryFile(dir=path.parent, suffix='.xlsx', delete=False) as stream:
            temporary = Path(stream.name)
        book.save(temporary)
        c.workbook_rows(temporary)
        temporary.replace(path)
        return True
    finally:
        book.close()


def verify_source():
    path=Path('/Users/ywh/Library/Application Support/com.TempoStorm.TheBazaar/prod/cache/GameData.db')
    if hashlib.sha256(path.read_bytes()).hexdigest()!=c.DB_SHA:raise ValueError('SOURCE_DB_SHA_MISMATCH')
    with sqlite3.connect(path.as_uri()+'?mode=ro',uri=True) as connection:
        card=json.loads(connection.execute('SELECT Data FROM cards WHERE Id=?',(c.SOURCE_UUID,)).fetchone()[0])
    assert card['Id']==c.SOURCE_UUID and card['Type']=='Item' and card['Size']=='Small'
    assert card['StartingTier']=='Silver' and card['Tags']==[] and card['HiddenTags']==['AmmoReference']
    assert card['Abilities']=={} and set(card['Tiers'])=={'Silver','Gold','Diamond'}
    for quality,amount in zip(('Silver','Gold','Diamond'),(1,2,3)):
        tier=card['Tiers'][quality]
        assert tier['Attributes']=={'Custom_0':amount} and tier['AbilityIds']==[] and tier['AuraIds']==['0']
    aura=card['Auras']['0']
    assert aura['ActiveIn']=='HandOnly' and aura['WorksIn']=='Anywhere' and aura['Prerequisites'] is None
    action=aura['Action']
    assert action['$type']=='TAuraActionCardModifyAttribute' and action['AttributeType']=='AmmoMax' and action['Operation']=='Add'
    assert action['Value']=={'$type':'TReferenceValueCardAttribute','AttributeType':'Custom_0','Target':{'$type':'TTargetCardSelf','Conditions':None},'DefaultValue':0.0,'Modifier':None}
    assert action['Target']=={'$type':'TTargetCardSection','TargetSection':'SelfHand','ExcludeSelf':False,'Conditions':{'$type':'TCardConditionalAttribute','Attribute':'AmmoMax','ComparisonOperator':'GreaterThan','ComparisonValue':{'$type':'TFixedValue','Value':0.0}}}
    # No raw card text/assets/ench payload is saved or returned.


def main():
    verify_source()
    rows=c.expected_rows()
    if c.WORKBOOK.exists():
        migrate_legacy_workbook(c.WORKBOOK, rows)
        assert c.workbook_rows(c.WORKBOOK)==c.canonical_rows(rows)
    else:
        book=openpyxl.Workbook();book.remove(book.active)
        for sheet,filename in c.SHEETS.items():
            target=book.create_sheet(sheet);headers=c.e.DOMAIN_HEADERS[filename]
            target.append(headers)
            for row in rows[filename]:target.append([row[k] for k in headers])
        c.WORKBOOK.parent.mkdir(parents=True,exist_ok=True)
        with tempfile.NamedTemporaryFile(dir=c.WORKBOOK.parent,suffix='.xlsx',delete=False) as stream:temporary=Path(stream.name)
        book.save(temporary);c.workbook_rows(temporary);temporary.replace(c.WORKBOOK)
    c.export_csv(c.WORKBOOK,c.CSV_DIR)
    print('Cannonball none source verified; isolated workbook and CSV generated; no acceptance claim')


if __name__=='__main__':main()
