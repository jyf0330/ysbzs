#!/usr/bin/env python3
"""Promote the locked Diving Helmet listener as fail-closed reference content."""

from pathlib import Path

import openpyxl

import export_original_pirate_content as e


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "xlsx" / "ysbzs_master.xlsx"
OLD_SNAPSHOT = "snapshot_original_pirate_bootstrap_v36"
NEW_SNAPSHOT = "snapshot_original_pirate_bootstrap_v37"
EXTERNAL_SNAPSHOT = "snapshot_vanessa_local_cache_25079259_db8914ab"
ITEM_ID = "item_bazaar_diving_helmet"
SOURCE_OBJECT_UUID = "fb6e6b16-d6d0-4493-ac3f-46c26afe6c51"
QUALITIES = (("gold", 50), ("diamond", 100))


def rows(sheet):
    header = [cell.value for cell in sheet[1]]
    return header, [dict(zip(header, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]


def append_dict(sheet, header, value):
    sheet.append([value.get(field, "") for field in header])


def main() -> None:
    workbook = openpyxl.load_workbook(MASTER)
    try:
        items = workbook["BZ_ITEMS"]
        item_header, item_rows = rows(items)
        if any(row["item_id"] == ITEM_ID for row in item_rows):
            raise ValueError("DIVING_HELMET_ITEM_ALREADY_PRESENT")
        for quality, _shield in QUALITIES:
            append_dict(items, item_header, {
                "item_id": ITEM_ID, "name_zh": "潜水头盔", "tags": "apparel,aquatic,tool",
                "slot_width": 2, "availability": "reference_battle_only",
                "base_quality": "gold", "quality": quality,
                "buy_price": "", "sell_price": "", "activation_mode": "passive",
                "cooldown_ticks": "", "crit_chance_bps": 0,
                "ammo_enabled": "false", "ammo_initial": 0, "ammo_maximum": 0,
                "item_skill_id": "skill_bazaar_diving_helmet_shield",
                "catalog_status": "formal",
            })

        effects = workbook["BZ_ITEM_EFFECTS"]
        effect_header, effect_rows = rows(effects)
        if any(row["item_id"] == ITEM_ID for row in effect_rows):
            raise ValueError("DIVING_HELMET_EFFECT_ALREADY_PRESENT")
        for quality, shield in QUALITIES:
            append_dict(effects, effect_header, {
                "effect_id": f"effect_bazaar_diving_helmet_{quality}_0", "item_id": ITEM_ID,
                "quality": quality, "item_skill_id": "skill_bazaar_diving_helmet_shield",
                "priority": 20, "trigger_event": "another_friendly_item_used",
                "condition_type": "source_item_has_any_tag", "condition_tags": "aquatic",
                "condition_source_relation": "any", "target_type": "owner_hero",
                "operation_type": "gain_shield", "amount": shield, "catalog_status": "formal",
                "source_ability_id": "0", "trigger_priority": "Medium", "effect_order": 0,
            })

        skills = workbook["BZ_ITEM_SKILLS"]
        skill_header, skill_rows = rows(skills)
        skill_id = "skill_bazaar_diving_helmet_shield"
        if any(row["item_skill_id"] == skill_id for row in skill_rows):
            raise ValueError("DIVING_HELMET_SKILL_ALREADY_PRESENT")
        append_dict(skills, skill_header, {
            "item_skill_id": skill_id, "name_zh": "深潜防护",
            "description_zh": "另一件己方水生物品使用后获得护盾；动态相邻标签与未证实同刻顺序不在本参考切片内。",
            "trigger_events": "another_friendly_item_used",
            "effect_ids": ",".join(f"effect_bazaar_diving_helmet_{q}_0" for q, _ in QUALITIES),
            "aura_ids": "", "catalog_status": "formal",
        })

        bindings = workbook["BZ_ITEM_SOURCE_BINDINGS"]
        binding_header, binding_rows = rows(bindings)
        if any(row["item_id"] == ITEM_ID for row in binding_rows):
            raise ValueError("DIVING_HELMET_BINDING_ALREADY_PRESENT")
        for quality, _shield in QUALITIES:
            append_dict(bindings, binding_header, {
                "item_id": ITEM_ID, "quality": quality, "enchantment_id": "none",
                "scope_id": "battle_profile", "source_snapshot_id": EXTERNAL_SNAPSHOT,
                "source_object_id": SOURCE_OBJECT_UUID,
            })
        for sheet_name in ("BZ_ITEM_SOURCE_BINDINGS", "BZ_HERO_SKILL_SOURCE_BINDINGS"):
            sheet = workbook[sheet_name]
            header = [cell.value for cell in sheet[1]]
            snapshot_col = header.index("source_snapshot_id") + 1
            for row_index in range(2, sheet.max_row + 1):
                if sheet.cell(row_index, snapshot_col).value == OLD_SNAPSHOT:
                    sheet.cell(row_index, snapshot_col, NEW_SNAPSHOT)

        ghost_snapshots = workbook["BZ_GHOST_SNAPSHOTS"]
        ghost_header = [cell.value for cell in ghost_snapshots[1]]
        revision_col = ghost_header.index("opponent_content_revision") + 1
        for row_index in range(2, ghost_snapshots.max_row + 1):
            if ghost_snapshots.cell(row_index, revision_col).value == "original-pirate-bootstrap-content-2026-09-05-v36":
                ghost_snapshots.cell(row_index, revision_col, "original-pirate-bootstrap-content-2026-09-05-v37")

        source = workbook["BZ_SOURCE_SNAPSHOT"]
        source_header = [cell.value for cell in source[1]]
        if source.cell(2, source_header.index("snapshot_id") + 1).value != OLD_SNAPSHOT:
            raise ValueError("SOURCE_SNAPSHOT_BASELINE_MISMATCH")
        source.cell(2, source_header.index("snapshot_id") + 1, NEW_SNAPSHOT)
        source.cell(2, source_header.index("source_revision") + 1, "original-pirate-bootstrap-source-2026-09-05-v37")
        source.cell(2, source_header.index("captured_on") + 1, "2026-09-05")
        scope_col = source_header.index("catalog_scope") + 1
        source.cell(2, scope_col, str(source.cell(2, scope_col).value) + "、Diving Helmet外部来源战斗档案引用")

        gameplay = workbook["BZ_GAMEPLAY"]
        gameplay_header = [cell.value for cell in gameplay[1]]
        updates = {
            "schema_version": "37", "runtime_schema_version": "35",
            "source_revision": "original-pirate-bootstrap-source-2026-09-05-v37",
            "bundle_revision": "original_pirate_bootstrap_bundle_v37",
            "content_revision": "original-pirate-bootstrap-content-2026-09-05-v37",
        }
        for row_index in range(2, gameplay.max_row + 1):
            for field, value in updates.items():
                gameplay.cell(row_index, gameplay_header.index(field) + 1, value)

        mapping_sheet = workbook["BZ_SOURCE_EFFECT_MAPPINGS"]
        mapping_header = [cell.value for cell in mapping_sheet[1]]
        common = {
            "mapping_id": "diving_helmet",
            "mapping_schema": "ysbzs.original-pirate-source-effect-mapping-candidate.v1",
            "mapping_schema_version": 1, "acceptance": "source_effect_mapping_only_not_complete_item",
            "source_data_commit": e.DIVING_HELMET_SOURCE_DATA_COMMIT,
            "mapping_sha256": e.DIVING_HELMET_MAPPING_SHA256,
            "provenance_sha256": e.DIVING_HELMET_PROVENANCE_SHA256,
            "source_db_sha256": "7d8df658ebce967edf59ab8d0c889fa266f56917b87336928694cdce54246ee9",
            "source_object_uuid": SOURCE_OBJECT_UUID, "source_internal_name": "Diving Helmet",
            "item_id": ITEM_ID, "effect_order": 0,
            "source_ability_id": "0", "source_trigger_type": "TTriggerOnItemUsed",
            "mapped_trigger_event": "friendly_item_used", "trigger_priority": "Medium",
            "target_type": "self_player", "target_condition_attribute": "Aquatic",
            "target_condition_operator": "Any", "subject_type": "self_hand_section",
            "subject_include_origin": "true", "operation_type": "gain_shield",
            "unknown_source_fields": "dynamicTagAuraApplicationTiming,dynamicTagAuraRemovalTiming,overlappingTagAuraReferenceCounting,sourceEventTagSnapshotPolicy,adjacencySnapshotAndMovementPolicy,disabledDestroyedTransformedAuraLifecycle,samePriorityCrossItemOrder,shieldResolutionOrder",
            "excluded_scopes": "enchantments,economy,acquisition,complete_initial_state,simultaneous_event_order,complete_run_state,top_three_identity",
            "catalog_status": "formal",
        }
        for quality, shield in QUALITIES:
            append_dict(mapping_sheet, mapping_header, dict(common, quality=quality, shield_apply_amount=shield))

        workbook.save(MASTER)
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
