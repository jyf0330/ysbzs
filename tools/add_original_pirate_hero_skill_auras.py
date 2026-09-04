"""Scoped workbook migration; verifies Circle identity without redistributing source payload."""
import hashlib,json,sqlite3,tempfile
from pathlib import Path
import openpyxl
import export_master_to_csv as m
import export_original_pirate_content as e
import bazaar_run_source_views as v

def main():
    db=Path('/Users/ywh/Library/Application Support/com.TempoStorm.TheBazaar/prod/cache/GameData.db')
    with db.open('rb') as f:
        if hashlib.file_digest(f,'sha256').hexdigest()!='7d8df658ebce967edf59ab8d0c889fa266f56917b87336928694cdce54246ee9':raise ValueError('DB_LOCK_MISMATCH')
    with sqlite3.connect(db.as_uri()+'?mode=ro',uri=True) as c:
        card=json.loads(c.execute('SELECT Data FROM cards WHERE Id=?',('9ec041be-6f89-4e95-963d-1deb7460e1d0',)).fetchone()[0])
    if card['Type']!='Skill' or sorted(card['Heroes'])!=['Karnok','Mak','Vanessa'] or card['SpawningEligibility']!='Always' or set(card['Tiers'])!={'Diamond'}:raise ValueError('CIRCLE_IDENTITY_MISMATCH')
    a=card['Auras']['0']['Action']
    if a['AttributeType']!='Lifesteal' or a['Operation']!='Add' or a['Value']!={'$type':'TFixedValue','Value':100.0}:raise ValueError('CIRCLE_VALUE_MISMATCH')
    if a['Target']!={'$type':'TTargetCardXMost','TargetSection':'SelfHand','TargetMode':'LeftMostCard','ExcludeSelf':True,'Conditions':{'$type':'TCardConditionalTag','Tags':['Weapon'],'Operator':'Any'}}:raise ValueError('CIRCLE_SELECTOR_MISMATCH')
    book=openpyxl.load_workbook(m.DEFAULT_MASTER)
    before={s.title:list(s.values) for s in book}
    gameplay=book['BZ_GAMEPLAY'];headers=[c.value for c in gameplay[1]]
    for field,value in [('rules_version',e.RULES_VERSION),('lifesteal_contract',e.LIFESTEAL_CONTRACT),('lifesteal_target_policy',e.LIFESTEAL_TARGET_POLICY)]:
        col=headers.index(field)+1
        for row in range(2,gameplay.max_row+1):gameplay.cell(row,col,value)
    skill=book['BZ_HERO_SKILLS']
    if 'aura_ids' not in [c.value for c in skill[1]]:skill.cell(1,skill.max_column+1,'aura_ids')
    aura_column=[c.value for c in skill[1]].index('aura_ids')+1
    for row in range(2,skill.max_row+1):
        if skill.cell(row,aura_column).value is None:skill.cell(row,aura_column,'')
    if 'BZ_HERO_SKILL_AURAS' not in book:
        book.create_sheet('BZ_HERO_SKILL_AURAS').append(e.DOMAIN_HEADERS['72_bz_hero_skill_auras.csv'])
    view_sheet=book[v.VIEW_SHEET]
    if v.CIRCLE_VIEW_ID not in [r[0] for r in view_sheet.values]:
        row=v.view_rows(m.RELOCKED_LOCAL_CACHE_SNAPSHOT_ID)[1];view_sheet.append([row[k] for k in v.VIEW_HEADERS])
    member_sheet=book[v.MEMBER_SHEET]
    if v.CIRCLE_VIEW_ID not in [r[0] for r in member_sheet.values]:
        row=v.circle_member();member_sheet.append([row[k] for k in v.MEMBER_HEADERS])
    owned={'BZ_GAMEPLAY','BZ_HERO_SKILLS','BZ_HERO_SKILL_AURAS',v.VIEW_SHEET,v.MEMBER_SHEET}
    with tempfile.NamedTemporaryFile(dir=m.DEFAULT_MASTER.parent,suffix='.xlsx',delete=False) as f:temporary=Path(f.name)
    book.save(temporary)
    check=openpyxl.load_workbook(temporary,read_only=True,data_only=False)
    if any(list(check[n].values)!=rows for n,rows in before.items() if n not in owned):raise ValueError('UNOWNED_SHEET_CHANGED')
    check.close();temporary.replace(m.DEFAULT_MASTER)
    tables=m.generated_original_pirate_tables(m.DEFAULT_MASTER)
    tables.update(m.generated_reference_source_tables(m.DEFAULT_MASTER))
    for name in ['44_bz_gameplay.csv','62_bz_hero_skills.csv','72_bz_hero_skill_auras.csv',v.VIEW_FILE,v.MEMBER_FILE]:
        rows,headers=tables[name];(e.DEFAULT_CSV_DIR/name).write_text(m.csv_text(rows,headers),encoding='utf-8')
    print('PASS scoped hero Aura workbook migration; Circle reference identity only')

if __name__=='__main__':main()
