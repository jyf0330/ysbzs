"""Add only the local skill identity declaration worksheet; preserve prior cells."""
import pathlib
import tempfile
import openpyxl
import export_master_to_csv as e
import original_pirate_source_binding as binding

def main():
    master=e.DEFAULT_MASTER
    table=e.generated_sheet_table(master,'BZ_HERO_SKILLS')[0]
    local=e.generated_sheet_table(master,'BZ_SOURCE_SNAPSHOT')[0][0]['snapshot_id']
    rows=sorted({(r['hero_skill_id'],r['quality'],'hero_skill_profile',local,r['hero_skill_id']) for r in table})
    name='BZ_HERO_SKILL_SOURCE_BINDINGS'
    workbook=openpyxl.load_workbook(master)
    before={s.title:list(s.values) for s in workbook}
    values=[tuple(binding.SKILL_HEADERS)]+rows
    if name in workbook:
        if list(workbook[name].values)!=values:raise ValueError('EXISTING_SKILL_SOURCE_SHEET_CONFLICT')
        print('PASS skill source worksheet already matches; no rewrite');return
    sheet=workbook.create_sheet(name)
    for row in values:sheet.append(row)
    with tempfile.NamedTemporaryFile(prefix='hero-sources-',suffix='.xlsx',dir=master.parent,delete=False) as f:
        temporary=pathlib.Path(f.name)
    workbook.save(temporary)
    check=openpyxl.load_workbook(temporary,read_only=True,data_only=False)
    assert all(list(check[n].values)==v for n,v in before.items()),'old worksheet changed'
    check.close();temporary.replace(master)
    print('PASS',len(rows),'local skill declarations;',len(before),'old worksheets unchanged')

if __name__=='__main__':main()
