#!/usr/bin/env python3
"""Promote the locked Pearl mapping as fail-closed reference-battle content."""

from pathlib import Path

import openpyxl

import export_original_pirate_content as e


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "xlsx" / "ysbzs_master.xlsx"
OLD_SNAPSHOT = "snapshot_original_pirate_bootstrap_v35"
NEW_SNAPSHOT = "snapshot_original_pirate_bootstrap_v36"
EXTERNAL_SNAPSHOT = "snapshot_vanessa_local_cache_25079259_db8914ab"
ITEM_ID = "item_bazaar_pearl"
SOURCE_OBJECT_UUID = "1312cf29-3dbb-446f-88b2-0d4999e68d78"
QUALITIES = (("bronze", 10), ("silver", 20), ("gold", 40), ("diamond", 80))


def rows(sheet):
    header = [cell.value for cell in sheet[1]]
    return header, [dict(zip(header, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]


def append_dict(sheet, header, value):
    sheet.append([value.get(field, "") for field in header])


def main() -> None:
    workbook = openpyxl.load_workbook(MASTER)
    try:
        items = workbook["BZ_ITEMS"]
        item_header, item_rows = rows(items)
        if any(row["item_id"] == ITEM_ID for row in item_rows):
            raise ValueError("PEARL_ITEM_ALREADY_PRESENT")
        for quality, _shield in QUALITIES:
            append_dict(items, item_header, {
                "item_id": ITEM_ID, "name_zh": "珍珠", "tags": "aquatic",
                "slot_width": 1, "availability": "reference_battle_only",
                "base_quality": "bronze", "quality": quality,
                "buy_price": "", "sell_price": "", "activation_mode": "cooldown",
                "cooldown_ticks": 100, "crit_chance_bps": 0,
                "ammo_enabled": "false", "ammo_initial": 0, "ammo_maximum": 0,
                "item_skill_id": "skill_bazaar_pearl_shield,skill_bazaar_pearl_charge",
                "catalog_status": "formal",
            })

        effects = workbook["BZ_ITEM_EFFECTS"]
        effect_header, effect_rows = rows(effects)
        if any(row["item_id"] == ITEM_ID for row in effect_rows):
            raise ValueError("PEARL_EFFECT_ALREADY_PRESENT")
        for quality, shield in QUALITIES:
            append_dict(effects, effect_header, {
                "effect_id": f"effect_bazaar_pearl_{quality}_0", "item_id": ITEM_ID,
                "quality": quality, "item_skill_id": "skill_bazaar_pearl_shield",
                "priority": 20, "trigger_event": "item_ready", "condition_type": "always",
                "condition_source_relation": "any", "target_type": "owner_hero",
                "operation_type": "gain_shield", "amount": shield, "catalog_status": "formal",
                "source_ability_id": "0", "trigger_priority": "Low", "effect_order": 0,
            })
            append_dict(effects, effect_header, {
                "effect_id": f"effect_bazaar_pearl_{quality}_1", "item_id": ITEM_ID,
                "quality": quality, "item_skill_id": "skill_bazaar_pearl_charge",
                "priority": 30, "trigger_event": "another_friendly_item_used",
                "condition_type": "source_item_has_any_tag", "condition_tags": "aquatic",
                "condition_source_relation": "any", "target_type": "self_item",
                "operation_type": "charge", "ticks": 20, "catalog_status": "formal",
                "source_ability_id": "1", "trigger_priority": "Low", "effect_order": 0,
            })

        skills = workbook["BZ_ITEM_SKILLS"]
        skill_header, skill_rows = rows(skills)
        for skill_id in ("skill_bazaar_pearl_shield", "skill_bazaar_pearl_charge"):
            if any(row["item_skill_id"] == skill_id for row in skill_rows):
                raise ValueError("PEARL_SKILL_ALREADY_PRESENT")
        append_dict(skills, skill_header, {
            "item_skill_id": "skill_bazaar_pearl_shield", "name_zh": "珍珠护盾",
            "description_zh": "珍珠使用时获得与品质相应的护盾。",
            "trigger_events": "item_ready",
            "effect_ids": ",".join(f"effect_bazaar_pearl_{q}_0" for q, _ in QUALITIES),
            "aura_ids": "", "catalog_status": "formal",
        })
        append_dict(skills, skill_header, {
            "item_skill_id": "skill_bazaar_pearl_charge", "name_zh": "水生共鸣",
            "description_zh": "另一件己方水生物品使用后，使珍珠推进一秒充能；若同刻触发未证实顺序或导致立即就绪则失败关闭。",
            "trigger_events": "another_friendly_item_used",
            "effect_ids": ",".join(f"effect_bazaar_pearl_{q}_1" for q, _ in QUALITIES),
            "aura_ids": "", "catalog_status": "formal",
        })

        bindings = workbook["BZ_ITEM_SOURCE_BINDINGS"]
        binding_header, binding_rows = rows(bindings)
        if any(row["item_id"] == ITEM_ID for row in binding_rows):
            raise ValueError("PEARL_BINDING_ALREADY_PRESENT")
        for quality, _shield in QUALITIES:
            append_dict(bindings, binding_header, {
                "item_id": ITEM_ID, "quality": quality, "enchantment_id": "none",
                "scope_id": "battle_profile", "source_snapshot_id": EXTERNAL_SNAPSHOT,
                "source_object_id": SOURCE_OBJECT_UUID,
            })
        for sheet_name in ("BZ_ITEM_SOURCE_BINDINGS", "BZ_HERO_SKILL_SOURCE_BINDINGS"):
            sheet = workbook[sheet_name]
            header = [cell.value for cell in sheet[1]]
            snapshot_col = header.index("source_snapshot_id") + 1
            for row_index in range(2, sheet.max_row + 1):
                if sheet.cell(row_index, snapshot_col).value == OLD_SNAPSHOT:
                    sheet.cell(row_index, snapshot_col, NEW_SNAPSHOT)

        ghost_snapshots = workbook["BZ_GHOST_SNAPSHOTS"]
        ghost_header = [cell.value for cell in ghost_snapshots[1]]
        opponent_revision_col = ghost_header.index("opponent_content_revision") + 1
        for row_index in range(2, ghost_snapshots.max_row + 1):
            if ghost_snapshots.cell(row_index, opponent_revision_col).value == "original-pirate-bootstrap-content-2026-09-05-v35":
                ghost_snapshots.cell(row_index, opponent_revision_col, "original-pirate-bootstrap-content-2026-09-05-v36")

        source = workbook["BZ_SOURCE_SNAPSHOT"]
        source_header = [cell.value for cell in source[1]]
        source_row = {field: source.cell(2, index + 1).value for index, field in enumerate(source_header)}
        if source_row["snapshot_id"] != OLD_SNAPSHOT:
            raise ValueError("SOURCE_SNAPSHOT_BASELINE_MISMATCH")
        source.cell(2, source_header.index("snapshot_id") + 1, NEW_SNAPSHOT)
        source.cell(2, source_header.index("source_revision") + 1, "original-pirate-bootstrap-source-2026-09-05-v36")
        source.cell(2, source_header.index("captured_on") + 1, "2026-09-05")
        scope_col = source_header.index("catalog_scope") + 1
        source.cell(2, scope_col, str(source.cell(2, scope_col).value) + "、Pearl外部来源战斗档案引用")

        gameplay = workbook["BZ_GAMEPLAY"]
        gameplay_header = [cell.value for cell in gameplay[1]]
        updates = {
            "schema_version": "36", "runtime_schema_version": "34",
            "source_revision": "original-pirate-bootstrap-source-2026-09-05-v36",
            "bundle_revision": "original_pirate_bootstrap_bundle_v36",
            "content_revision": "original-pirate-bootstrap-content-2026-09-05-v36",
        }
        for row_index in range(2, gameplay.max_row + 1):
            for field, value in updates.items():
                gameplay.cell(row_index, gameplay_header.index(field) + 1, value)

        mapping_sheet = workbook["BZ_SOURCE_EFFECT_MAPPINGS"]
        mapping_header = [cell.value for cell in mapping_sheet[1]]
        if "shield_apply_amount" not in mapping_header:
            insert_at = mapping_header.index("source_ability_id") + 1
            mapping_sheet.insert_cols(insert_at)
            mapping_sheet.cell(1, insert_at, "shield_apply_amount")
            mapping_header = [cell.value for cell in mapping_sheet[1]]
        common = {
            "mapping_id": "pearl", "mapping_schema": "ysbzs.original-pirate-source-effect-mapping-candidate.v1",
            "mapping_schema_version": 1, "acceptance": "source_effect_mapping_only_not_complete_item",
            "source_data_commit": e.PEARL_SOURCE_DATA_COMMIT,
            "mapping_sha256": e.PEARL_MAPPING_SHA256, "provenance_sha256": e.PEARL_PROVENANCE_SHA256,
            "source_db_sha256": "7d8df658ebce967edf59ab8d0c889fa266f56917b87336928694cdce54246ee9",
            "source_object_uuid": SOURCE_OBJECT_UUID, "source_internal_name": "Pearl",
            "item_id": ITEM_ID, "cooldown_max_milliseconds": 5000, "multicast": 1,
            "charge_amount_milliseconds": 1000, "charge_targets": 1, "effect_order": 0,
            "unknown_source_fields": "initialCooldownProgress,samePriorityCrossItemOrder,samePriorityAbilityOrder,cardFiredToItemUsedDispatchOrder,chargeCausedReadyReentrancy,shieldAndChargeSameTimestampOrder,multicastCardFiredPolicy,completeCritAndStatusInitialState",
            "excluded_scopes": "enchantment_execution,economy,acquisition,complete_initial_state,simultaneous_ready_order,complete_run_state,top_three_identity",
            "catalog_status": "formal",
        }
        for quality, shield in QUALITIES:
            append_dict(mapping_sheet, mapping_header, dict(common, quality=quality,
                shield_apply_amount=shield, source_ability_id="0",
                source_trigger_type="TTriggerOnCardFired", mapped_trigger_event="item_ready",
                trigger_priority="Low", target_type="self_player", operation_type="gain_shield"))
            append_dict(mapping_sheet, mapping_header, dict(common, quality=quality,
                shield_apply_amount=shield, source_ability_id="1",
                source_trigger_type="TTriggerOnItemUsed", mapped_trigger_event="another_friendly_item_used",
                trigger_priority="Low", target_type="self", subject_type="self_hand_section",
                subject_include_origin="false", target_condition_attribute="Aquatic",
                target_condition_operator="Any", operation_type="charge", ticks=20))

        workbook.save(MASTER)
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
