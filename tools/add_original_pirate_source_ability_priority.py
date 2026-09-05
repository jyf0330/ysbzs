#!/usr/bin/env python3
"""Append optional source Ability columns without changing existing workbook cells."""
from pathlib import Path
import argparse
import hashlib
import json
import sqlite3
import tempfile

import openpyxl

import export_master_to_csv as master_export
import export_original_pirate_content as content_export


SHEET = "BZ_ITEM_EFFECTS"
CSV_NAME = "47_bz_item_effects.csv"
NEW_FIELDS = ("source_ability_id", "trigger_priority", "effect_order")
DB_SHA256 = "7d8df658ebce967edf59ab8d0c889fa266f56917b87336928694cdce54246ee9"
FIXTURE = content_export.ROOT / "tests/support/original_pirate_source_ability_priority_fixture.json"


def verify_locked_source(db_path: Path) -> None:
    with db_path.open("rb") as stream:
        if hashlib.file_digest(stream, "sha256").hexdigest() != DB_SHA256:
            raise ValueError("SOURCE_ABILITY_DB_LOCK_MISMATCH")
    fixture = json.loads(FIXTURE.read_text(encoding="utf-8"))
    if fixture["sourceDbSha256"] != DB_SHA256:
        raise ValueError("SOURCE_ABILITY_FIXTURE_DB_LOCK_MISMATCH")
    expected = {
        (row["sourceObjectUuid"], row["sourceAbilityId"]): row["triggerPriority"]
        for row in fixture["priorityExamples"]
    }
    rifle_uuid = fixture["rifle"]["sourceObjectUuid"]
    with sqlite3.connect(db_path.resolve().as_uri() + "?mode=ro", uri=True) as connection:
        for (uuid, ability_id), priority in expected.items():
            row = connection.execute("SELECT Data FROM cards WHERE Id=?", (uuid,)).fetchone()
            if row is None or json.loads(row[0])["Abilities"][ability_id]["Priority"] != priority:
                raise ValueError("SOURCE_ABILITY_PRIORITY_IDENTITY_MISMATCH")
        rifle_row = connection.execute("SELECT Data FROM cards WHERE Id=?", (rifle_uuid,)).fetchone()
    rifle = json.loads(rifle_row[0]) if rifle_row else {}
    actual_rifle = []
    for ability_id, ability in rifle.get("Abilities", {}).items():
        action = ability.get("Action", {})
        actual_rifle.append({
            "sourceAbilityId": ability_id,
            "triggerType": ability.get("Trigger", {}).get("$type"),
            "triggerPriority": ability.get("Priority"),
            "actionType": action.get("$type") if isinstance(action, dict) else None,
            "effectOrder": 0,
        })
    if actual_rifle != fixture["rifle"]["abilities"]:
        raise ValueError("SOURCE_ABILITY_RIFLE_IDENTITY_MISMATCH")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append source Ability columns after verifying a locked GameData.db snapshot."
    )
    parser.add_argument(
        "--db",
        required=True,
        type=Path,
        help="Path to the locked GameData.db used by the source identity fixture.",
    )
    return parser.parse_args()


def main(db_path: Path) -> None:
    verify_locked_source(db_path)
    master = master_export.DEFAULT_MASTER
    workbook = openpyxl.load_workbook(master)
    before = {sheet.title: list(sheet.values) for sheet in workbook}
    sheet = workbook[SHEET]
    headers = [cell.value for cell in sheet[1]]
    expected_old = content_export.DOMAIN_HEADERS[CSV_NAME][:-len(NEW_FIELDS)]
    if headers == content_export.DOMAIN_HEADERS[CSV_NAME]:
        workbook.close()
        tables = master_export.generated_original_pirate_tables(master)
        rows, exported_headers = tables[CSV_NAME]
        (content_export.DEFAULT_CSV_DIR / CSV_NAME).write_text(
            master_export.csv_text(rows, exported_headers), encoding="utf-8"
        )
        print("PASS source Ability columns already present; workbook unchanged")
        return
    if headers != expected_old:
        raise ValueError("ITEM_EFFECT_HEADERS_UNEXPECTED")
    for field in NEW_FIELDS:
        sheet.cell(1, sheet.max_column + 1, field)
    with tempfile.NamedTemporaryFile(
        prefix=master.stem + ".source-ability-priority.", suffix=".xlsx",
        dir=master.parent, delete=False,
    ) as handle:
        temporary = Path(handle.name)
    workbook.save(temporary)
    workbook.close()
    check = openpyxl.load_workbook(temporary, read_only=True, data_only=False)
    for name, old_rows in before.items():
        current = list(check[name].values)
        if name != SHEET and current != old_rows:
            raise ValueError("UNOWNED_SHEET_CHANGED:" + name)
        if name == SHEET:
            if current[0] != tuple(content_export.DOMAIN_HEADERS[CSV_NAME]):
                raise ValueError("ITEM_EFFECT_HEADERS_NOT_APPENDED")
            if any(row[:len(old_rows[0])] != old for row, old in zip(current[1:], old_rows[1:])):
                raise ValueError("ITEM_EFFECT_OLD_CELL_CHANGED")
    check.close()
    master_export.generated_original_pirate_tables(temporary)
    temporary.replace(master)
    tables = master_export.generated_original_pirate_tables(master)
    rows, exported_headers = tables[CSV_NAME]
    (content_export.DEFAULT_CSV_DIR / CSV_NAME).write_text(
        master_export.csv_text(rows, exported_headers), encoding="utf-8"
    )
    print(f"PASS appended {len(NEW_FIELDS)} columns; all existing cells unchanged")


if __name__ == "__main__":
    main(parse_args().db)
