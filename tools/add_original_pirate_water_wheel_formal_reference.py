#!/usr/bin/env python3
"""Promote the locked Water Wheel mapping as a formal reference-battle-only profile."""

from pathlib import Path

import openpyxl

import export_original_pirate_content as e


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "xlsx" / "ysbzs_master.xlsx"
OLD_SNAPSHOT = "snapshot_original_pirate_bootstrap_v34"
NEW_SNAPSHOT = "snapshot_original_pirate_bootstrap_v35"
EXTERNAL_SNAPSHOT = "snapshot_vanessa_local_cache_25079259_db8914ab"
ITEM_ID = "item_bazaar_water_wheel"
QUALITIES = (("silver", 160), ("gold", 140), ("diamond", 120))


def rows(sheet):
    header = [cell.value for cell in sheet[1]]
    return header, [dict(zip(header, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]


def append_dict(sheet, header, value):
    sheet.append([value.get(field, "") for field in header])


def main() -> None:
    workbook = openpyxl.load_workbook(MASTER)
    try:
        items = workbook["BZ_ITEMS"]
        item_header = [cell.value for cell in items[1]]
        if "availability" not in item_header:
            slot_index = item_header.index("slot_width") + 2
            items.insert_cols(slot_index)
            items.cell(1, slot_index, "availability")
            item_header = [cell.value for cell in items[1]]
            availability_col = item_header.index("availability") + 1
            for row_index in range(2, items.max_row + 1):
                items.cell(row_index, availability_col, "run_acquirable")
        item_header, item_rows = rows(items)
        if any(row["item_id"] == ITEM_ID for row in item_rows):
            raise ValueError("WATER_WHEEL_ITEM_ALREADY_PRESENT")
        for quality, cooldown in QUALITIES:
            append_dict(items, item_header, {
                "item_id": ITEM_ID, "name_zh": "水车", "tags": "aquatic",
                "slot_width": 3, "availability": "reference_battle_only",
                "base_quality": "silver", "quality": quality,
                "buy_price": "", "sell_price": "", "activation_mode": "cooldown",
                "cooldown_ticks": cooldown, "crit_chance_bps": 0,
                "ammo_enabled": "false", "ammo_initial": 0, "ammo_maximum": 0,
                "item_skill_id": "skill_bazaar_water_wheel_haste,skill_bazaar_water_wheel_charge",
                "catalog_status": "formal",
            })

        effects = workbook["BZ_ITEM_EFFECTS"]
        effect_header, effect_rows = rows(effects)
        if any(row["item_id"] == ITEM_ID for row in effect_rows):
            raise ValueError("WATER_WHEEL_EFFECT_ALREADY_PRESENT")
        for quality, _cooldown in QUALITIES:
            append_dict(effects, effect_header, {
                "effect_id": f"effect_bazaar_water_wheel_{quality}_0", "item_id": ITEM_ID,
                "quality": quality, "item_skill_id": "skill_bazaar_water_wheel_haste",
                "priority": 20, "trigger_event": "item_ready", "condition_type": "always",
                "condition_source_relation": "any", "target_type": "all_other_friendly_active_clock_items",
                "target_exclude_self": "true", "operation_type": "apply_status",
                "status": "haste", "ticks": 40, "catalog_status": "formal",
                "source_ability_id": "0", "trigger_priority": "High", "effect_order": 0,
            })
            append_dict(effects, effect_header, {
                "effect_id": f"effect_bazaar_water_wheel_{quality}_1", "item_id": ITEM_ID,
                "quality": quality, "item_skill_id": "skill_bazaar_water_wheel_charge",
                "priority": 30, "trigger_event": "another_friendly_item_used", "condition_type": "always",
                "condition_source_relation": "adjacent", "target_type": "self_item",
                "operation_type": "charge", "ticks": 40, "catalog_status": "formal",
                "source_ability_id": "1", "trigger_priority": "Medium", "effect_order": 0,
            })

        skills = workbook["BZ_ITEM_SKILLS"]
        skill_header, skill_rows = rows(skills)
        for skill_id in ("skill_bazaar_water_wheel_haste", "skill_bazaar_water_wheel_charge"):
            if any(row["item_skill_id"] == skill_id for row in skill_rows):
                raise ValueError("WATER_WHEEL_SKILL_ALREADY_PRESENT")
        append_dict(skills, skill_header, {
            "item_skill_id": "skill_bazaar_water_wheel_haste", "name_zh": "水车急流",
            "description_zh": "使用后，使己方其他具有有效冷却时钟的物品获得两秒急速；重复急速规则未知时失败关闭。",
            "trigger_events": "item_ready",
            "effect_ids": ",".join(f"effect_bazaar_water_wheel_{q}_0" for q, _ in QUALITIES),
            "aura_ids": "", "catalog_status": "formal",
        })
        append_dict(skills, skill_header, {
            "item_skill_id": "skill_bazaar_water_wheel_charge", "name_zh": "邻轮推进",
            "description_zh": "相邻的另一件己方物品使用后，使水车推进两秒充能。",
            "trigger_events": "another_friendly_item_used",
            "effect_ids": ",".join(f"effect_bazaar_water_wheel_{q}_1" for q, _ in QUALITIES),
            "aura_ids": "", "catalog_status": "formal",
        })

        bindings = workbook["BZ_ITEM_SOURCE_BINDINGS"]
        binding_header, binding_rows = rows(bindings)
        if any(row["item_id"] == ITEM_ID for row in binding_rows):
            raise ValueError("WATER_WHEEL_BINDING_ALREADY_PRESENT")
        for quality, _cooldown in QUALITIES:
            append_dict(bindings, binding_header, {
                "item_id": ITEM_ID, "quality": quality, "enchantment_id": "none",
                "scope_id": "battle_profile", "source_snapshot_id": EXTERNAL_SNAPSHOT,
                "source_object_id": "d8106a24-647f-40c6-8587-22f977931d76",
            })
        for sheet_name in ("BZ_ITEM_SOURCE_BINDINGS", "BZ_HERO_SKILL_SOURCE_BINDINGS"):
            sheet = workbook[sheet_name]
            header = [cell.value for cell in sheet[1]]
            snapshot_col = header.index("source_snapshot_id") + 1
            for row_index in range(2, sheet.max_row + 1):
                if sheet.cell(row_index, snapshot_col).value == OLD_SNAPSHOT:
                    sheet.cell(row_index, snapshot_col, NEW_SNAPSHOT)

        ghost_snapshots = workbook["BZ_GHOST_SNAPSHOTS"]
        ghost_header = [cell.value for cell in ghost_snapshots[1]]
        opponent_revision_col = ghost_header.index("opponent_content_revision") + 1
        for row_index in range(2, ghost_snapshots.max_row + 1):
            if ghost_snapshots.cell(row_index, opponent_revision_col).value == "original-pirate-bootstrap-content-2026-09-04-v34":
                ghost_snapshots.cell(row_index, opponent_revision_col, "original-pirate-bootstrap-content-2026-09-05-v35")

        source = workbook["BZ_SOURCE_SNAPSHOT"]
        source_header = [cell.value for cell in source[1]]
        source_row = {field: source.cell(2, index + 1).value for index, field in enumerate(source_header)}
        if source_row["snapshot_id"] != OLD_SNAPSHOT:
            raise ValueError("SOURCE_SNAPSHOT_BASELINE_MISMATCH")
        source.cell(2, source_header.index("snapshot_id") + 1, NEW_SNAPSHOT)
        source.cell(2, source_header.index("source_revision") + 1, "original-pirate-bootstrap-source-2026-09-05-v35")
        source.cell(2, source_header.index("captured_on") + 1, "2026-09-05")
        scope_col = source_header.index("catalog_scope") + 1
        source.cell(2, scope_col, str(source.cell(2, scope_col).value) + "、Water Wheel外部来源战斗档案引用")

        gameplay = workbook["BZ_GAMEPLAY"]
        gameplay_header = [cell.value for cell in gameplay[1]]
        updates = {
            "schema_version": "35", "runtime_schema_version": "33",
            "rules_version": "ysbzs.original-pirate-rules.2026-09-05-v34",
            "source_revision": "original-pirate-bootstrap-source-2026-09-05-v35",
            "bundle_revision": "original_pirate_bootstrap_bundle_v35",
            "content_revision": "original-pirate-bootstrap-content-2026-09-05-v35",
        }
        for row_index in range(2, gameplay.max_row + 1):
            for field, value in updates.items():
                gameplay.cell(row_index, gameplay_header.index(field) + 1, value)

        if "BZ_SOURCE_EFFECT_MAPPINGS" in workbook.sheetnames:
            raise ValueError("SOURCE_EFFECT_MAPPING_SHEET_ALREADY_PRESENT")
        mapping_sheet = workbook.create_sheet("BZ_SOURCE_EFFECT_MAPPINGS")
        mapping_header = e.DOMAIN_HEADERS["73_bz_source_effect_mappings.csv"]
        mapping_sheet.append(mapping_header)
        common = {
            "mapping_id": "water_wheel", "mapping_schema": "ysbzs.original-pirate-source-effect-mapping-candidate.v1",
            "mapping_schema_version": 1, "acceptance": "source_effect_mapping_only_not_complete_item",
            "source_data_commit": e.WATER_WHEEL_SOURCE_DATA_COMMIT,
            "mapping_sha256": e.WATER_WHEEL_MAPPING_SHA256, "provenance_sha256": e.WATER_WHEEL_PROVENANCE_SHA256,
            "source_db_sha256": "7d8df658ebce967edf59ab8d0c889fa266f56917b87336928694cdce54246ee9",
            "source_object_uuid": "d8106a24-647f-40c6-8587-22f977931d76", "source_internal_name": "Water Wheel",
            "item_id": ITEM_ID, "multicast": 1, "haste_amount_milliseconds": 2000,
            "charge_amount_milliseconds": 2000, "charge_targets": 1, "effect_order": 0,
            "unknown_source_fields": "initialCooldownProgress,hasteReapplicationPolicy,same_priority_tie_break",
            "excluded_scopes": "enchantments,economy,acquisition,complete_initial_state,simultaneous_ready_order,complete_run_state,top_three_identity",
            "catalog_status": "formal",
        }
        for quality, cooldown_ticks in QUALITIES:
            cooldown_ms = cooldown_ticks * 50
            append_dict(mapping_sheet, mapping_header, dict(common, quality=quality,
                cooldown_max_milliseconds=cooldown_ms, source_ability_id="0",
                source_trigger_type="TTriggerOnCardFired", mapped_trigger_event="item_ready",
                trigger_priority="High", target_type="self_hand_section", target_exclude_self="true",
                target_condition_attribute="CooldownMax", target_condition_operator="GreaterThan",
                target_condition_value=0, operation_type="apply_status", status="haste", ticks=40))
            append_dict(mapping_sheet, mapping_header, dict(common, quality=quality,
                cooldown_max_milliseconds=cooldown_ms, source_ability_id="1",
                source_trigger_type="TTriggerOnItemUsed", mapped_trigger_event="another_friendly_item_used",
                trigger_priority="Medium", target_type="self", subject_type="self_positional",
                subject_target_mode="Neighbor", subject_include_origin="false",
                operation_type="charge", ticks=40))

        workbook.save(MASTER)
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
