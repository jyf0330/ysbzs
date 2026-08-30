#!/usr/bin/env python3
"""Build the first planner-facing master workbook from current CSV truth."""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "data" / "csv"
OUT = ROOT / "xlsx" / "ysbzs_master.xlsx"


def read_csv(name):
    with (CSV_DIR / name).open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_csv_rows(name):
    with (CSV_DIR / name).open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def split_pool_count(expr):
    raw = str(expr or "").strip()
    if "-" not in raw:
        return raw, ""
    left, right = raw.rsplit("-", 1)
    if right.strip().isdigit() and left.strip():
        return left.strip(), right.strip()
    return raw, ""


def add_sheet(wb, title, headers, rows, widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, header in enumerate(headers, start=1):
        width = (widths or {}).get(header)
        if width is None:
            sample = [str(header)] + [str(row.get(header, "")) for row in rows[:30]]
            width = min(max(max(len(s) for s in sample) + 2, 10), 34)
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def add_domain_sheet(wb, title, sections):
    ws = wb.create_sheet(title)
    section_fill = PatternFill("solid", fgColor="7030A0")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_font = Font(color="FFFFFF", bold=True)
    header_font = Font(color="FFFFFF", bold=True)
    for csv_name, description in sections:
        ws.append(["#csv", csv_name, description])
        section_row = ws.max_row
        for cell in ws[section_row]:
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        rows = read_csv_rows(csv_name)
        if rows:
            ws.append(rows[0])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in rows[1:]:
                ws.append(row)
        ws.append([])
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        sample = [str(ws.cell(r, col_idx).value or "") for r in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(max((len(s) for s in sample), default=8) + 2, 10), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def main():
    pets = read_csv("01_pets.csv")
    monsters = read_csv("02_monster_templates.csv")
    waves = read_csv("03_monster_waves.csv")
    shop = read_csv("06_shop_rewards.csv")
    shop_stores = read_csv("30_shop_stores.csv")
    start_choices = read_csv("43_start_choices.csv")
    shop_by_pet = {r.get("宠物ID", ""): r for r in shop}
    monster_by_pet = {r.get("宠物ID", ""): r for r in monsters}

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    readme = wb.create_sheet("README")
    readme_rows = [
        ("用途", "这是人类策划入口；程序完整数据仍输出到 data/csv/*.csv。"),
        ("日常维护", "优先改 PETS / SHOP_STORES / WAVES / SHOP_ITEMS；机制、品质、形状、试炼在合并分区表里维护。"),
        ("工作表", "README / PETS / SHOP_STORES / WAVES / SHOP_ITEMS / MECHANICS_QUALITY / SHAPES_TRIALS，共 7 张可见表。"),
        ("导出", "npm run data:export"),
        ("校验", "npm run check:csv"),
        ("价格规则", "商店公开价按品质统一导出：青铜=2、白银=4、黄金=6、钻石=8。"),
        ("边界", "自动列、程序冗余列不放进日常主表；需要完整好读版时运行 npm run data:workbook。"),
    ]
    readme.append(["项目", "说明"])
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 16
    readme.column_dimensions["B"].width = 86
    readme.freeze_panes = "A2"
    for cell in readme[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)

    add_sheet(
        wb,
        "PETS",
        [
            "pet_id",
            "name",
            "element",
            "tier",
            "role",
            "shop_store_ids",
            "hp",
            "atk",
            "shield",
            "action",
            "mechanism_id",
            "shape_id",
            "note",
            "enemy_move_range",
            "enemy_attack_count",
            "skill_ids",
        ],
        [
            {
                "pet_id": r.get("宠物ID", ""),
                "name": r.get("名称", ""),
                "element": "、".join([x for x in [r.get("元素", ""), r.get("副属", "")] if x]),
                "tier": r.get("品质", ""),
                "role": r.get("定位", ""),
                "shop_store_ids": shop_by_pet.get(r.get("宠物ID", ""), {}).get("商店池(自动)", ""),
                "hp": r.get("HP", ""),
                "atk": r.get("攻", ""),
                "shield": r.get("盾", ""),
                "action": r.get("行动", ""),
                "mechanism_id": r.get("机制ID", ""),
                "shape_id": r.get("形状", ""),
                "note": r.get("备注", ""),
                "enemy_move_range": monster_by_pet.get(r.get("宠物ID", ""), {}).get("移动力", r.get("行动", "")),
                "enemy_attack_count": monster_by_pet.get(r.get("宠物ID", ""), {}).get("攻击次数", r.get("行动", "")),
                "skill_ids": r.get("技能序列", ""),
            }
            for r in pets
        ],
        widths={
            "pet_id": 12,
            "shop_store_ids": 38,
            "mechanism_id": 26,
            "shape_id": 18,
            "note": 26,
            "enemy_move_range": 20,
            "enemy_attack_count": 20,
            "skill_ids": 52,
        },
    )

    add_sheet(
        wb,
        "SHOP_STORES",
        ["shop_store_id", "name", "store_type", "tags", "default_slots", "unlock_day", "price_rule", "status", "note"],
        [
            {
                "shop_store_id": r.get("shop_store_id", ""),
                "name": r.get("name", ""),
                "store_type": r.get("store_type", ""),
                "tags": r.get("tags", ""),
                "default_slots": r.get("default_slots", ""),
                "unlock_day": r.get("unlock_day", ""),
                "price_rule": r.get("price_rule", ""),
                "status": r.get("status", ""),
                "note": r.get("note", ""),
            }
            for r in shop_stores
        ],
        widths={"shop_store_id": 22, "tags": 22, "price_rule": 14, "note": 34},
    )

    add_sheet(
        wb,
        "WAVES",
        ["wave_id", "day", "period", "round", "enemy_pool", "count", "quality_weights", "target_threat", "design_goal"],
        [
            {
                "wave_id": r.get("波次ID", ""),
                "day": r.get("天数", ""),
                "period": r.get("时段", ""),
                "round": r.get("回合", ""),
                "enemy_pool": split_pool_count(r.get("宠物池-数量", ""))[0],
                "count": split_pool_count(r.get("宠物池-数量", ""))[1],
                "quality_weights": r.get("品质权重", ""),
                "target_threat": r.get("本行威胁(当前计算值)", ""),
                "design_goal": r.get("填写说明", ""),
            }
            for r in waves
        ],
        widths={"wave_id": 22, "enemy_pool": 28, "quality_weights": 18, "design_goal": 36, "note": 26},
    )

    add_sheet(
        wb,
        "SHOP_ITEMS",
        ["pet_id", "unlock_day", "tier_pool", "shop_weight", "reward_weight", "note"],
        [
            {
                "pet_id": r.get("宠物ID", ""),
                "unlock_day": r.get("解锁日", ""),
                "tier_pool": r.get("池档", ""),
                "shop_weight": r.get("夜市权重", ""),
                "reward_weight": r.get("奖励权重", ""),
                "note": r.get("备注", ""),
            }
            for r in shop
        ],
        widths={"pet_id": 12, "note": 28},
    )

    add_domain_sheet(wb, "MECHANICS_QUALITY", [
        ("04_mechanisms.csv", "机制注册表：机制 ID、触发、效果、接入状态。"),
        ("28_quality_growth.csv", "品质成长数值。"),
        ("29_quality_upgrades.csv", "品质升级质变。"),
        ("31_battle_rules.csv", "正式双人战斗可调规则。"),
    ])

    add_domain_sheet(wb, "SHAPES_TRIALS", [
        ("27_shape_catalog.csv", "19 个战斗形状目录；offsets/grid 会影响运行时攻击范围和 UI 展示。"),
        ("15_summon_trial_questions.csv", "召唤试炼题库。"),
        ("16_trial_action_plan.csv", "试炼行动脚本。"),
        ("17_trial_victory_rules.csv", "试炼胜负规则。"),
        ("36_skill_catalog.csv", "8 技能 Type Object；effects_json 依次执行技能效果。"),
        ("37_trait_catalog.csv", "宠物特性 Type Object；按 skill/combo hook 修饰效果。"),
        ("38_skill_combo_catalog.csv", "有序技能组合 Type Object；按相邻技能标签顺序匹配。"),
    ])

    add_domain_sheet(wb, "ATTRIBUTES_EFFECTS", [
        ("39_stat_catalog.csv", "通用宠物属性目录；整数/千分比、默认值、上下限和叠加规则。"),
        ("40_status_catalog.csv", "运行状态 Type Object；叠层、持续回合和通用 modifier effects。"),
    ])

    add_sheet(
        wb,
        "START_CHOICES",
        [
            "start_choice_id", "name", "description", "display_order", "coins_delta",
            "free_rolls_delta", "pet_id", "pet_quality", "run_health_delta",
            "run_max_health_delta", "status", "source", "note",
        ],
        start_choices,
        widths={"start_choice_id": 27, "description": 46, "note": 42},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
