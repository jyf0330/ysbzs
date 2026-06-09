#!/usr/bin/env python3
"""Rebuild planner-readable workbook sheets from data/csv.

This is the opposite side of the thin-master workflow:
- xlsx/ysbzs_master.xlsx is the daily human-editing entry.
- data/csv/*.csv is the complete program input.
- xlsx/ysbzs_v1_linked_data_tables.xlsx remains the broad readable workbook
  for review, browsing, and handoff.
"""

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "data" / "csv"
DEFAULT_TARGET = ROOT / "xlsx" / "ysbzs_v1_linked_data_tables.xlsx"

CSV_TO_SHEET = [
    ("00_maintenance_guide.csv", "00_读我_怎么维护"),
    ("01_pets.csv", "01_宠物主表_好读版"),
    ("02_monster_templates.csv", "02_怪物模板_好读版"),
    ("03_monster_waves.csv", "03_怪物波次_好读版"),
    ("04_mechanisms.csv", "04_机制词条库_60条"),
    ("05_events.csv", "05_事件主表_好读版"),
    ("06_shop_rewards.csv", "06_商店奖励池_好读版"),
    ("07_relic_blessings.csv", "07_遗物祝福_好读版"),
    ("08_action_shapes.csv", "08_形状行动槽_好读版"),
    ("09_cross_validation.csv", "09_跨表校验"),
    ("10_initial_roster.csv", "10_初始阵容"),
    ("11_hero_domains.csv", "11_英雄领域_联动版"),
    ("12_element_reactions.csv", "12_元素反应_联动版"),
    ("13_day7_beast_trial.csv", "13_第7天兽群试炼_联动版"),
    ("14_quality_multipliers.csv", "14_品质进阶倍率"),
    ("15_summon_trial_questions.csv", "15_召唤试炼题库"),
]


def read_csv_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def clear_sheet(ws):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def column_width(values):
    max_len = max((len(str(v or "")) for v in values), default=8)
    return min(max(max_len + 2, 10), 34)


def write_sheet(ws, rows):
    clear_sheet(ws)
    for row in rows:
        ws.append(row)
    if not rows:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        sample = [ws.cell(r, col_idx).value for r in range(1, min(ws.max_row, 40) + 1)]
        ws.column_dimensions[get_column_letter(col_idx)].width = column_width(sample)


def load_or_create(path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def main():
    parser = argparse.ArgumentParser(description="Rebuild planner-readable workbook sheets from data/csv")
    parser.add_argument("--target", default=str(DEFAULT_TARGET))
    parser.add_argument("--csv-dir", default=str(CSV_DIR))
    parser.add_argument("--only-existing", action="store_true", help="Update only sheets already present in the target workbook")
    args = parser.parse_args()

    target = Path(args.target)
    csv_dir = Path(args.csv_dir)
    wb = load_or_create(target)
    updated = []
    for csv_name, sheet_name in CSV_TO_SHEET:
        csv_path = csv_dir / csv_name
        if not csv_path.exists():
            continue
        if args.only_existing and sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        write_sheet(ws, read_csv_rows(csv_path))
        updated.append(sheet_name)

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f"updated {len(updated)} readable sheets -> {target}")


if __name__ == "__main__":
    main()
