"""Real diamond Circle of Life skill identity + static Aura, synthetic acquisition host."""
import argparse,copy,csv,hashlib,json,tempfile
from pathlib import Path
import openpyxl
import export_original_pirate_content as e
import export_original_pirate_cannonball_candidate as shared
import bazaar_run_source_views as view

SKILL_ID='hero_skill_bazaar_circle_of_life'
UUID='9ec041be-6f89-4e95-963d-1deb7460e1d0'
WORKBOOK=e.ROOT/'xlsx/candidates/original_pirate_circle_of_life.xlsx'
CSV_DIR=e.ROOT/'data/candidates/original_pirate/circle_of_life'
SHEETS={'BZ_HERO_SKILLS':'62_bz_hero_skills.csv','BZ_HERO_SKILL_TRAINERS':'64_bz_hero_skill_trainers.csv','BZ_HERO_SKILL_OFFERS':'65_bz_hero_skill_offers.csv','BZ_HERO_SKILL_SOURCE_BINDINGS':'71_bz_hero_skill_source_bindings.csv','BZ_HERO_SKILL_AURAS':'72_bz_hero_skill_auras.csv'}

def expected_rows():
    tables={n:[] for n in SHEETS.values()}
    def add(n,**kw):tables[n].append({k:str(kw.get(k,'')) for k in e.DOMAIN_HEADERS[n]})
    add('62_bz_hero_skills.csv',hero_skill_id=SKILL_ID,hero_id='hero_mistwake_captain',quality='diamond',name_zh='生命循环（原版钻石候选）',description_zh='已装备英雄技能为己方最左武器提供吸血；仅静态候选。',effect_description_zh='己方最左武器增加百分之一百吸血，其他结算规则仍依项目宿主。',priority=30,trigger_event='none',reentrant='false',max_triggers_per_battle=0,catalog_status='formal',aura_ids='hero_aura_circle_of_life_diamond')
    add('64_bz_hero_skill_trainers.csv',trainer_id='trainer_candidate_circle_of_life',hero_id='hero_mistwake_captain',stall_id='stall_mistwake',name_zh='生命循环测试席',description_zh='仅测试取得路径，非原作技能商店。',offer_slots=1,catalog_status='formal')
    add('65_bz_hero_skill_offers.csv',offer_id='offer_candidate_circle_of_life',trainer_id='trainer_candidate_circle_of_life',hero_skill_id=SKILL_ID,action_type='learn',to_quality='diamond',price_currency='gold',price_amount=1,from_day=1,to_day=10,offer_order=1,name_zh='学习生命循环候选',description_zh='测试宿主以一金币取得钻石候选技能，不代表原作报价。',catalog_status='formal')
    add('71_bz_hero_skill_source_bindings.csv',hero_skill_id=SKILL_ID,quality='diamond',scope_id='hero_skill_profile',source_snapshot_id=view.CIRCLE_VIEW_ID,source_object_id=UUID)
    add('72_bz_hero_skill_auras.csv',aura_id='hero_aura_circle_of_life_diamond',hero_skill_id=SKILL_ID,quality='diamond',priority=0,target_type='leftmost_friendly_item_with_any_tag',target_tags='weapon',operation_type='grant_lifesteal_bps',lifesteal_bps=10000,catalog_status='formal')
    return tables

def validate_rows(tables):
    if tables!=expected_rows():raise ValueError('CIRCLE_CANDIDATE_TRANSLATION_LOCK_MISMATCH')

def initialize():
    expected=expected_rows()
    if not WORKBOOK.exists():
        book=openpyxl.Workbook();book.remove(book.active)
        for s,n in SHEETS.items():
            ws=book.create_sheet(s);headers=e.DOMAIN_HEADERS[n];ws.append(headers)
            for row in expected[n]:ws.append([row[k] for k in headers])
        WORKBOOK.parent.mkdir(parents=True,exist_ok=True)
        with tempfile.NamedTemporaryFile(dir=WORKBOOK.parent,suffix='.xlsx',delete=False) as f:p=Path(f.name)
        book.save(p);p.replace(WORKBOOK)
    rows=workbook_rows()
    for n,values in rows.items():shared.atomic_write(CSV_DIR/n,shared.csv_text(n,values))

def workbook_rows():
    book=openpyxl.load_workbook(WORKBOOK,read_only=True,data_only=False)
    try:
        if set(book.sheetnames)!=set(SHEETS):raise ValueError('CIRCLE_CANDIDATE_SHEETS_INVALID')
        result={}
        for s,n in SHEETS.items():
            values=list(book[s].values);headers=e.DOMAIN_HEADERS[n]
            if list(values[0])!=headers:raise ValueError('CIRCLE_CANDIDATE_HEADERS_INVALID')
            result[n]=[dict(zip(headers,['' if v is None else str(v) for v in r])) for r in values[1:]]
        validate_rows(result);return result
    finally:book.close()

def build_candidate():
    tables=workbook_rows()
    for n,rows in tables.items():
        if (CSV_DIR/n).read_text(encoding='utf-8')!=shared.csv_text(n,rows):raise ValueError('CIRCLE_CSV_STALE')
    # Reuse the exact baseline lock, not the Cannonball item overlay.
    base_sha=hashlib.sha256(json.dumps({n:hashlib.sha256((e.DEFAULT_CSV_DIR/n).read_bytes()).hexdigest() for n in e.DOMAIN_HEADERS},sort_keys=True,separators=(',',':')).encode()).hexdigest()
    base,_=e.build_exports(e.DEFAULT_CSV_DIR)
    if base_sha!=shared.BASE_CSV_SHA or base['runtimeBundle']['bundleHash']!=shared.BASE_HASH:raise ValueError('CIRCLE_BASE_LOCK_MISMATCH')
    merged=e._read_domains(e.DEFAULT_CSV_DIR)
    for n,rows in tables.items():merged[n].extend(copy.deepcopy(rows))
    p=e.ContentAssembler(merged,e.DEFAULT_CSV_DIR).build();e.validate_package(p)
    d=e._build_display_directory(merged,p)
    provenance={'schema':'ysbzs.original-pirate-circle-of-life-candidate.v1','acceptance':'not_original_game_acceptance','baseBundleHash':shared.BASE_HASH,'bundleHash':p['runtimeBundle']['bundleHash'],'sourceDbSha256':shared.DB_SHA,'sourceUuid':UUID,'heroSkillId':SKILL_ID,'declaredQualities':['diamond'],'sourceViewId':view.CIRCLE_VIEW_ID,'sourceMechanism':'equipped skill: leftmost friendly Weapon gains 100% Lifesteal','syntheticHost':['trainer and diamond learn offer price 1 gold','hero owner identity','definition and Aura priority'],'limitations':['not original acquisition or full Run initial state','project Lifesteal HP-damage basis, cap and healing/cleanse rules are not original-game verified','no dynamic destruction/transformation evidence','no complete build popularity or six-match acceptance']}
    return p,d,provenance

def main():
    parser=argparse.ArgumentParser(description=__doc__);parser.add_argument('--initialize',action='store_true');parser.add_argument('--out-dir',type=Path)
    args=parser.parse_args()
    if args.initialize:initialize()
    if args.out_dir:
        if args.out_dir.resolve()==e.DEFAULT_CSV_DIR.resolve() or any(k in args.out_dir.resolve().parts for k in ('generated','gameplay')):raise ValueError('CIRCLE_OUTPUT_MUST_BE_ISOLATED')
        values=build_candidate();outputs={n:json.dumps(v,ensure_ascii=False,indent=2)+'\n' for n,v in zip(('content.json','display.zh-CN.json','provenance.json'),values)}
        for n,t in outputs.items():
            p=args.out_dir/n
            if p.exists() and p.read_text(encoding='utf-8')!=t:raise ValueError('CIRCLE_OUTPUT_EXISTS_DIFFERENT')
        for n,t in outputs.items():shared.atomic_write(args.out_dir/n,t)
        print(values[0]['runtimeBundle']['bundleHash'])

if __name__=='__main__':main()
